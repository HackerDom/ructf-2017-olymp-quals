import openpyxl
import PIL.Image as Image
import PIL
import openpyxl.styles as styles
def create_document():
    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    image = Image.open("task_pic.png")
    pixel_map = image.load()
    width, height = image.size
    for i in range(1,width):
        for j in range(1,height):
            cell = work_sheet.cell(row = j , column = i)
            pixel_color  =  pixel_map[i,j]
            color = "FF{0:02X}{1:02X}{2:02X}".format(*pixel_color)
            cell.fill = styles.PatternFill(fill_type="solid", start_color=color, end_color=color)
            if (i * height + j) % 1000 == 0 :
                progress_bar((i* height + j) / (width * height))
    work_book.save("task3.xlsx")

def progress_bar(state):
    st1 = int(state * 100)
    st2 = int(state * 10000) % 100
    val = (25 - (st1 // 4))
    start = "=" *(25 - val)
    end = ' '* val
    bar = "[[{}{}]({}.{}%)".format(start, end, st1, st2)
    print('\r',bar, end='')
create_document()

